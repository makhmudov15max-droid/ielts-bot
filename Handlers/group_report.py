import os, io, re, json, logging, asyncio
from html import unescape
from datetime import datetime, timezone, timedelta
from aiogram import Router, F, types
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
import config

logger = logging.getLogger(__name__)

report_router = Router()

from utils.group_comments_db import get_comment, set_comment, delete_comment, get_all_comments

try:
    ADMIN_ID = int(config.ADMIN_ID)
except:
    ADMIN_ID = 6500594896


# ================= STATES =================
class ReportStates(StatesGroup):
    waiting_for_report_choice = State()
    waiting_for_teacher_choice = State()
    waiting_for_comment_input = State()
    waiting_for_comment_delete_confirm = State()
    waiting_for_cashbox_detail = State()
    waiting_for_teacher_name_add = State()
    waiting_for_teacher_name_remove = State()
    # Qarzdorlik foizi -> maosh hisoblash
    waiting_for_salary_choice = State()
    waiting_for_salary_hours = State()
    waiting_for_salary_days = State()
    waiting_for_salary_cover = State()


# ================= GLOBAL O'ZGARUVCHI =================
_USERS_ROLES = None

# Cache: {key: (result, timestamp)}
_REPORT_CACHE = {}
_CACHE_TTL = 60  # soniya

# DRUJBA_TEACHERS — JSON faylda saqlanadi
_TEACHERS_FILE = os.path.join(os.path.dirname(__file__), "..", "data", "drujba_teachers.json")

def _load_teachers():
    """DRUJBA_TEACHERS ro'yxatini JSON fayldan o'qish"""
    default_teachers = [
        "Sardor Komilov",
        "Adxambek Ismoilov",
        "Otabek Mirhamidov",
        "Ahmadali Turgunov",
        "Obidjon Rustamov",
        "Odiljon Jaloliddinov",
        "Ibrohim Aliyev",
        "Xurshid Hazratqulov",
        "Farangiz Elamanova",
        "Sevinch Ibrohimova",
        "Nilufar Karimova",
    ]
    try:
        fpath = os.path.abspath(_TEACHERS_FILE)
        if os.path.exists(fpath):
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list) and len(data) > 0:
                return data
        # Fayl yo'q yoki noto'g'ri — yaratib qo'yamiz
        os.makedirs(os.path.dirname(fpath), exist_ok=True)
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(default_teachers, f, ensure_ascii=False, indent=2)
        return default_teachers
    except Exception as e:
        logger.warning(f"_load_teachers error: {e}")
        return default_teachers

def _save_teachers(teachers: list):
    """DRUJBA_TEACHERS ro'yxatini JSON faylga yozish"""
    try:
        fpath = os.path.abspath(_TEACHERS_FILE)
        os.makedirs(os.path.dirname(fpath), exist_ok=True)
        with open(fpath, "w", encoding="utf-8") as f:
            json.dump(teachers, f, ensure_ascii=False, indent=2)
        # Keshni tozalash
        _REPORT_CACHE.pop("finance", None)
        return True
    except Exception as e:
        logger.error(f"_save_teachers error: {e}")
        return False

def set_users_roles(users_roles):
    global _USERS_ROLES
    _USERS_ROLES = users_roles


async def is_admin(user_id: int) -> bool:
    if _USERS_ROLES:
        user_info = _USERS_ROLES.get(str(user_id))
        if user_info:
            role = user_info.get("role")
            if role in ["Owner", "Manager", "Manager Assistant", "Head Admin"]:
                return True
    from utils.users_db import get_user_role
    role = await get_user_role(str(user_id))
    return role in ["Owner", "Manager", "Manager Assistant", "Head Admin"]


# ================= LMS CONFIG =================
LMS_BASE = "https://main.ieltszoneapp.uz"
LMS_EMAIL = config.LMS_EMAIL if hasattr(config, 'LMS_EMAIL') else "makhmudov15max@gmail.com"
LMS_KEY = config.LMS_KEY if hasattr(config, 'LMS_KEY') else os.getenv("LMS_KEY", "Mahmudov02")

DRUJBA_BRANCH_ID = 3
IELTS_COURSE_IDS = {7, 8, 9, 10, 12, 15}  # Novice, Standard, Expert, Intensive, Practice, Speaking
UZ_TZ = timezone(timedelta(hours=5))

# Cache
_lms_session = None
_teacher_map = {}
_course_map = {}
_cached_groups = None
_cached_groups_time = 0
_cached_cashboxes = None
_cached_cashboxes_time = 0
CACHE_TTL = 60  # sekund


def _get_lms_session():
    """LMS sessiyasini yaratish yoki qayta ishlatish"""
    global _lms_session, _teacher_map, _course_map

    import requests
    from urllib.parse import unquote
    from html import unescape

    # Agar sessiya mavjud bo'lsa, tekshirib ko'rish
    if _lms_session:
        try:
            r = _lms_session.get(f"{LMS_BASE}/admin/dashboard", timeout=10)
            if r.status_code == 200 and "data-page" in r.text:
                return _lms_session
        except:
            pass

    # Yangi sessiya
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "X-Requested-With": "XMLHttpRequest",
    })
    s.get(f"{LMS_BASE}/sanctum/csrf-cookie")
    xsrf = s.cookies.get("XSRF-TOKEN", "")
    s.headers["X-XSRF-TOKEN"] = unquote(xsrf)
    s.headers["Content-Type"] = "application/json"
    r = s.post(f"{LMS_BASE}/admin/login", json={"email": LMS_EMAIL, "password": LMS_KEY})
    logger.info(f"LMS login status: {r.status_code}")

    # Get teacher + course maps
    if not _teacher_map or not _course_map:
        try:
            r = s.get(f"{LMS_BASE}/admin/unassessed-groups?per_page=1")
            match = re.search(r'data-page="([^\"]*)"', r.text)
            if match:
                dp = json.loads(unescape(match.group(1)))
                for c in dp["props"]["courseOptions"]:
                    _course_map[c["id"]] = c["name"].get("uz", c["name"].get("en", ""))
                for t in dp["props"]["teacherOptions"]:
                    _teacher_map[t["id"]] = f"{t.get('first_name','')} {t.get('last_name','')}".strip()

            # Also get from calculated-salaries (all employees)
            r = s.get(f"{LMS_BASE}/admin/calculated-salaries?per_page=200")
            match = re.search(r'data-page="([^\"]*)"', r.text)
            if match:
                dp = json.loads(unescape(match.group(1)))
                for emp in dp["props"]["employees"]:
                    if emp["id"] not in _teacher_map:
                        _teacher_map[emp["id"]] = f"{emp.get('first_name','')} {emp.get('last_name','')}".strip()
        except Exception as e:
            logger.warning(f"LMS maps fetch error: {e}")

    _lms_session = s
    return s


def get_all_groups():
    """LMS export orqali barcha Drujba IELTS guruhlarni olish (aktiv + kutilayotgan)"""
    global _cached_groups, _cached_groups_time
    import time

    # Cache ishlatish — 60 soniya davomida qayta yuklanmaydi
    now = time.time()
    if _cached_groups is not None and (now - _cached_groups_time) < CACHE_TTL:
        logger.info(f"LMS cache: {len(_cached_groups)} groups ({(now - _cached_groups_time):.0f}s old)")
        return _cached_groups

    import openpyxl

    try:
        s = _get_lms_session()
        r = s.get(f"{LMS_BASE}/admin/groups/export", timeout=30)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active

        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        name_col = headers.index("Ism") + 1
        start_col = headers.index("Guruhning boshlanish sanasi") + 1
        end_col = headers.index("craftable-pro.Group End Date") + 1
        status_col = headers.index("Status") + 1
        course_col = headers.index("Kurs Id") + 1
        teacher_col = headers.index("craftable-pro.Teacher Id") + 1
        branch_col = headers.index("Fillial Id") + 1

        today = datetime.now(UZ_TZ).date()
        groups = []

        for row_idx in range(2, ws.max_row + 1):
            bid = ws.cell(row_idx, branch_col).value
            if bid != DRUJBA_BRANCH_ID:
                continue

            cid = ws.cell(row_idx, course_col).value
            if cid not in IELTS_COURSE_IDS:
                continue

            status = ws.cell(row_idx, status_col).value
            if status not in {1, 2}:  # 1=draft/kutilmoqda, 2=aktiv
                continue

            # Teacher ID dan nom olish
            tid = ws.cell(row_idx, teacher_col).value
            tname = _teacher_map.get(tid, f"ID#{tid}") if tid else ""

            # Draft guruhlarda teacher=None bo'lsa, nomidan ajratish
            if not tname or tname.startswith("ID#"):
                gname = str(ws.cell(row_idx, name_col).value or "")
                for _, map_name in sorted(_teacher_map.items(), key=lambda x: -len(x[1])):
                    if map_name.lower() in gname.lower():
                        tname = map_name
                        break

            # Agar teacher aniqlanmasa, guruhni tashlab ketamiz
            if not tname:
                continue

            cname = _course_map.get(cid, f"ID#{cid}")

            start_str = str(ws.cell(row_idx, start_col).value or "")
            start_fmt = ""
            if start_str and start_str != "None":
                try:
                    start_fmt = datetime.strptime(start_str[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
                except:
                    pass

            # End date va days_left
            end_str = str(ws.cell(row_idx, end_col).value or "")
            end_fmt = ""
            days_left = 999
            if end_str and end_str != "None":
                try:
                    end_fmt = datetime.strptime(end_str[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
                    end_date = datetime.strptime(end_str[:10], "%Y-%m-%d").date()
                    days_left = (end_date - today).days
                except:
                    pass

            groups.append({
                "name": f"{ws.cell(row_idx, name_col).value}",
                "start": start_fmt,
                "end": end_fmt,
                "days_left": days_left,
                "status": status,
                "course": cname,
                "teacher": tname,
            })

        _cached_groups = groups
        _cached_groups_time = now
        logger.info(f"LMS: {len(groups)} Drujba IELTS groups")
        return groups

    except Exception as e:
        logger.error(f"LMS export error: {e}")
        return []


# ================= LMS HISOBOT HANDLERLARI (Reply Keyboard) =================
@report_router.message(ReportStates.waiting_for_report_choice, F.text == "📋 Finishing Groups")
async def lms_finishing_handler(message: types.Message, state: FSMContext):
    """LMS -> Finishing Groups"""
    groups = await asyncio.to_thread(get_all_groups)
    finishing = [g for g in groups if 0 <= g["days_left"] <= 14]
    if not finishing:
        await message.answer("📭 Finishing groups topilmadi.")
        return
    text = f"📋 <b>Finishing Groups</b> | Jami: {len(finishing)}\n\n"
    for g in sorted(finishing, key=lambda x: x["days_left"]):
        status = "⚪" if g["status"] == 1 else "🟢"
        text += f"{status} {g['name']} — {g['course']} | {g['teacher']}\n"
        text += f"    Tugash: {g['end']} ({g['days_left']} kun)\n\n"
    await message.answer(text, parse_mode="HTML")


@report_router.message(ReportStates.waiting_for_report_choice, F.text == "⏳ Waiting Groups")
async def show_waiting_groups(message: types.Message, state: FSMContext):
    if not await is_admin(message.from_user.id):
        await message.answer("⚠️ Bu buyruq faqat administrator va owner uchun!")
        return

    await message.answer("⏳ Kutilayotgan guruhlar yuklanmoqda...")
    groups = await asyncio.to_thread(_get_waiting_groups)
    all_comments = await get_all_comments()

    if not groups:
        await message.answer("📭 Hozircha kutilayotgan guruhlar yo'q.")
        return

    await state.update_data(waiting_groups=groups, report_type="waiting")

    report, inline_kb = _render_waiting_groups_report(groups, all_comments)
    await message.answer(report, parse_mode="HTML", reply_markup=inline_kb)


@report_router.message(ReportStates.waiting_for_report_choice, F.text == "👨🏻‍🏫 Ustoz bo'yicha")
async def show_teachers_list(message: types.Message, state: FSMContext):
    teachers = await asyncio.to_thread(get_unique_teachers)

    if not teachers:
        await message.answer("📭 LMSda o'qituvchilar topilmadi.")
        return

    await state.set_state(ReportStates.waiting_for_teacher_choice)

    inline_kb = []
    for t in teachers:
        inline_kb.append([types.InlineKeyboardButton(
            text=f"👨🏻‍🏫 {t}",
            callback_data=f"teachergroups_{t}",
        )])

    inline_kb.append([types.InlineKeyboardButton(
        text="🏠 Bosh sahifa",
        callback_data="teachergroups_cancel",
    )])

    await message.answer(
        text=f"👨🏻‍🏫 <b>Ustozni tanlang:</b>\n\nJami {len(teachers)} ta o'qituvchi",
        parse_mode="HTML",
        reply_markup=types.InlineKeyboardMarkup(inline_keyboard=inline_kb),
    )


@report_router.message(ReportStates.waiting_for_report_choice, F.text == "📋 Dars Jadval")
async def export_schedule_to_sheets(message: types.Message, state: FSMContext):
    """LMS dan dars jadvalini olib Google Sheets ga yozadi."""
    await message.answer("⏳ LMS dan dars jadvali yuklanmoqda...\n\nBu biroz vaqt olishi mumkin (30-60 soniya).")

    from utils.sheets_export import write_schedule_to_sheets
    result = await write_schedule_to_sheets()
    await message.answer(result, parse_mode="HTML")


@report_router.message(ReportStates.waiting_for_report_choice, F.text == "💰 Finance Report")
async def show_finance_report(message: types.Message, state: FSMContext):
    """Finance report faqat Owner uchun"""
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    if role != "Owner":
        await message.answer("⚠️ Bu buyruq faqat Owner uchun!")
        return

    msg = await message.answer("⏳ Finance report tayyorlanmoqda...")

    try:
        # Cache check
        now = datetime.now(timezone.utc).timestamp()
        cached = _REPORT_CACHE.get("finance")
        if cached and now - cached[1] < _CACHE_TTL:
            await msg.edit_text(cached[0], parse_mode="HTML")
            return

        s = _get_lms_session()

        DRUJBA_TEACHERS = _load_teachers()

        r = s.get(f"{LMS_BASE}/admin/calculated-salaries?per_page=200", timeout=15)
        match = re.search(r'data-page="([^"]*)"', r.text)
        api_balances = {}
        if match:
            dp = json.loads(unescape(match.group(1)))
            for emp in dp["props"]["employees"]:
                name = f"{emp.get('first_name','')} {emp.get('last_name','')}".strip()
                eb = emp.get("employee_balance", "0")
                try:
                    eb = float(eb)
                except:
                    eb = 0
                api_balances[name] = eb

        def _match_teacher(expected_name, all_balances):
            parts = expected_name.lower().split()
            for api_name, bal in all_balances.items():
                api_lower = api_name.lower()
                if all(p in api_lower for p in parts):
                    return bal, api_name
            return 0, expected_name

        text = "💰 <b>DRUJBA — USTOZLAR BALANSI</b>\n\n"
        total_balance = 0
        teacher_count = 0
        for teacher in DRUJBA_TEACHERS:
            bal, matched_name = _match_teacher(teacher, api_balances)
            total_balance += bal
            teacher_count += 1
            if bal >= 0:
                text += f"{teacher_count}. 👨🏻‍🏫 {teacher}\n   💰 {int(bal)} so'm\n\n"
            else:
                text += f"{teacher_count}. 👨🏻‍🏫 {teacher}\n   🔴 {int(bal)} so'm\n\n"

        text += f"━━━━━━━━━━━━━━━━\n"
        text += f"📊 <b>Jami:</b> {teacher_count} ta ustoz\n"
        text += f"💵 <b>Umumiy balans:</b> {int(total_balance)} so'm"

        # Cache
        _REPORT_CACHE["finance"] = (text, datetime.now(timezone.utc).timestamp())

        await msg.edit_text(text, parse_mode="HTML")

    except Exception as e:
        logger.error(f"Finance report error: {e}")
        await msg.edit_text(f"⚠️ Xatolik yuz berdi: {e}")


@report_router.message(ReportStates.waiting_for_report_choice, F.text == "🏦 Cashbox")
async def show_cashboxes(message: types.Message, state: FSMContext):
    """Barcha cashboxlarni listing qiladi"""
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    if role != "Owner":
        await message.answer("⚠️ Bu buyruq faqat Owner uchun!")
        return

    msg = await message.answer("⏳ Cashboxlar yuklanmoqda...")

    try:
        s = _get_lms_session()
        r = s.get(f"{LMS_BASE}/admin/cashboxes", timeout=15)
        match = re.search(r'data-page="([^"]*)"', r.text)
        if not match:
            await msg.edit_text("⚠️ Cashbox ma'lumotlari topilmadi.")
            return

        dp = json.loads(unescape(match.group(1)))
        cashboxes = dp["props"]["cashboxes"]

        if not cashboxes:
            await msg.edit_text("📭 Hech qanday cashbox topilmadi.")
            return

        drujba_cbs = [cb for cb in cashboxes if cb.get("branch", {}).get("en") == "Drujba filial"]

        text = "🏦 <b>DRUJBA — CASHBOXLAR</b>\n\n"
        from_keyboard = []

        for idx, cb in enumerate(drujba_cbs, 1):
            bal = cb.get("balance", {}) if isinstance(cb.get("balance"), dict) else {}
            total = sum(float(v or 0) for v in bal.values())
            text += f"{idx}. <b>{cb['name']}</b> — 💰 {int(total)} so'm\n"
            from_keyboard.append([types.InlineKeyboardButton(text=f"{cb['name']} — {int(total)} so'm", callback_data=f"cb_{cb['id']}")])

        from_keyboard.append([types.InlineKeyboardButton(text="🏠 Bosh sahifa", callback_data="cb_home")])
        inline_kb = types.InlineKeyboardMarkup(inline_keyboard=from_keyboard)

        await state.set_state(ReportStates.waiting_for_cashbox_detail)
        await msg.edit_text(text, parse_mode="HTML", reply_markup=inline_kb)

    except Exception as e:
        logger.error(f"Cashbox list error: {e}")
        await msg.edit_text(f"⚠️ Xatolik yuz berdi: {e}")


@report_router.message(ReportStates.waiting_for_report_choice, F.text == "🏠 Bosh sahifa")
async def lms_home_handler(message: types.Message, state: FSMContext):
    """LMS -> Bosh sahifa"""
    await state.clear()
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    from Keyboards.main_menu import get_main_menu
    await message.answer("🏠 <b>Bosh sahifa</b>", parse_mode="HTML", reply_markup=get_main_menu(role))


@report_router.message(ReportStates.waiting_for_report_choice, F.text == "⬅️ Ortga")
async def lms_back_handler(message: types.Message, state: FSMContext):
    """LMS submenu -> LMS asosiy menyu"""
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
    if role == "Kassir":
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="💲 Debtors")],
                [KeyboardButton(text="🏠 Bosh sahifa")],
            ],
            resize_keyboard=True,
        )
    else:
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="📂 Groups"), KeyboardButton(text="💰 Finance")],
                [KeyboardButton(text="💲 Debtors")],
                [KeyboardButton(text="🏠 Bosh sahifa")],
            ],
            resize_keyboard=True,
        )
    await message.answer(
        "📑 <b>LMS Panel</b>\n\nQuyidagi bo'limlardan birini tanlang:",
        parse_mode="HTML",
        reply_markup=kb,
    )


@report_router.message(F.text == "🌐 LMS")
async def lms_main_handler(message: types.Message, state: FSMContext):
    """LMS tugmasi bosilganda — Reply keyboard panelni ko'rsatish"""
    # Ruxsatsiz foydalanuvchilarga butunlay jim javob (Admin ham kira olmaydi)
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", None) if _USERS_ROLES else None
    if role != "Kassir" and not await is_admin(message.from_user.id):
        return

    await state.set_state(ReportStates.waiting_for_report_choice)
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton

    if role == "Kassir":
        # Kassirga LMS ichidan faqat '💲 Debtors' ko'rinadi (Debs ％ uchun)
        base_buttons = [
            [KeyboardButton(text="💲 Debtors")],
            [KeyboardButton(text="🏠 Bosh sahifa")],
        ]
    else:
        # Asosiy LMS menyusi
        base_buttons = [
            [KeyboardButton(text="📂 Groups"), KeyboardButton(text="💰 Finance")],
            [KeyboardButton(text="💲 Debtors")],
            [KeyboardButton(text="🏠 Bosh sahifa")],
        ]

    await message.answer(
        "📑 <b>LMS Panel</b>\n\nQuyidagi bo'limlardan birini tanlang:",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(keyboard=base_buttons, resize_keyboard=True),
    )


# ================= GROUPS SUBMENU =================

@report_router.message(ReportStates.waiting_for_report_choice, F.text == "📂 Groups")
async def groups_submenu(message: types.Message, state: FSMContext):
    """Groups submanyusi: Waiting, Finishing, Dars Jadval, Ustoz bo'yicha"""
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📊 Finishing Groups"), KeyboardButton(text="⏳ Waiting Groups")],
            [KeyboardButton(text="👨🏻‍🏫 Ustoz bo'yicha"), KeyboardButton(text="📋 Dars Jadval")],
            [KeyboardButton(text="⬅️ Ortga")],
        ],
        resize_keyboard=True,
    )
    await message.answer("📂 <b>Groups</b> — kerakli bo'limni tanlang:", parse_mode="HTML", reply_markup=kb)


# ================= FINANCE SUBMENU =================

@report_router.message(ReportStates.waiting_for_report_choice, F.text == "💰 Finance")
async def finance_submenu(message: types.Message, state: FSMContext):
    """Finance submanyusi: Finance Report, Cashbox, Ustozlarni boshqarish"""
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
    if role == "Owner":
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="💰 Finance Report"), KeyboardButton(text="🏦 Cashbox")],
                [KeyboardButton(text="👨🏻‍🏫 Ustozlarni boshqarish")],
                [KeyboardButton(text="⬅️ Ortga")],
            ],
            resize_keyboard=True,
        )
    else:
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="💰 Finance Report"), KeyboardButton(text="🏦 Cashbox")],
                [KeyboardButton(text="⬅️ Ortga")],
            ],
            resize_keyboard=True,
        )
    await message.answer("💰 <b>Finance</b> — kerakli bo'limni tanlang:", parse_mode="HTML", reply_markup=kb)


# ================= YORDAMCHI FUNKSIYALAR =================

def get_unique_teachers():
    """LMS guruhlardan unikal ustozlar ro'yxatini qaytaradi"""
    groups = get_all_groups()
    teachers = set()
    for g in groups:
        if g["teacher"]:
            teachers.add(g["teacher"])
    return sorted(teachers)


def _get_waiting_groups():
    """Drujba status=1 (kutilayotgan) guruhlarni xona va o'quvchilar soni bilan qaytaradi"""
    import openpyxl
    from html import unescape
    s = _get_lms_session()
    BASE = LMS_BASE

    r = s.get(f"{BASE}/admin/groups/export", timeout=30)
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active

    h = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    id_col = h.index("Id") + 1
    name_col = h.index("Ism") + 1
    status_col = h.index("Status") + 1
    teacher_col = h.index("craftable-pro.Teacher Id") + 1
    course_col = h.index("Kurs Id") + 1
    branch_col = h.index("Fillial Id") + 1

    waiting = []
    for row_idx in range(2, ws.max_row + 1):
        bid = ws.cell(row_idx, branch_col).value
        status = ws.cell(row_idx, status_col).value
        if bid == 3 and status == 1:
            gid = ws.cell(row_idx, id_col).value
            gname = str(ws.cell(row_idx, name_col).value or "")
            tid = ws.cell(row_idx, teacher_col).value
            cid = ws.cell(row_idx, course_col).value
            tname = _teacher_map.get(tid, "❌ Tayinlanmagan") if tid else "❌ Tayinlanmagan"
            cname = _course_map.get(cid, f"ID#{cid}")

            try:
                r2 = s.get(f"{BASE}/admin/groups/{gid}", timeout=15)
                m2 = re.search(r'data-page="([^"]*)"', r2.text)
                if m2:
                    dp = json.loads(unescape(m2.group(1)))
                    gd = dp["props"]["group"]

                    rooms = gd.get("rooms", [])
                    room_name = rooms[0]["name"] if rooms else "❌"
                    capacity = rooms[0]["capacity"] if rooms else 20

                    students = gd.get("students", gd.get("newly_added_trial_frozen_active_failed_students", []))
                    # trial holati (status 1, 8) oldin hisoblanar edi — Sales bo'limi o'chirilgani uchun olib tashlandi
                    active = sum(1 for s in students if s.get("pivot", {}).get("status") == 6)
                    frozen = sum(1 for s in students if s.get("pivot", {}).get("status") == 9)

                    waiting.append({
                        "name": gname,
                        "teacher": tname,
                        "level": cname,
                        "room": room_name,
                        "capacity": capacity,
                        "active": active,
                        "frozen": frozen,
                    })
            except Exception as e:
                logger.warning(f"Waiting group #{gid} fetch error: {e}")
                waiting.append({
                    "name": gname,
                    "teacher": tname,
                    "level": cname,
                    "room": "❌",
                    "capacity": 0,
                    "active": 0,
                    "frozen": 0,
                })

    return waiting


def _render_waiting_groups_report(groups: list, all_comments: dict):
    text = f"⏳ <b>WAITING GROUPS</b> — {len(groups)} ta guruh\n\n"

    num_rows = []
    row = []
    for idx, g in enumerate(groups):
        has_comment = bool(all_comments.get(g["name"], ""))
        label = f"{idx + 1}{'📝' if has_comment else ''}"
        row.append(types.InlineKeyboardButton(
            text=label,
            callback_data=f"wgp_{idx}"
        ))
        if len(row) >= 5:
            num_rows.append(row)
            row = []
    if row:
        num_rows.append(row)

    inline_kb = types.InlineKeyboardMarkup(inline_keyboard=num_rows) if num_rows else None

    global_idx = 0
    by_teacher = {}
    for g in groups:
        by_teacher.setdefault(g["teacher"], []).append(g)

    for teacher, gs in sorted(by_teacher.items()):
        text += f"👨🏻‍🏫 <b>{teacher}</b>\n"
        for g in gs:
            global_idx += 1
            comment = all_comments.get(g["name"], "")
            comment_line = f" 📝" if comment else ""
            text += (
                f"   <b>{global_idx}.</b> 📚 {g['name']} — {g['level']}{comment_line}\n"
                f"       👥 {g['active']} + {g['frozen']} / {'?' if g['capacity'] == 0 else g['capacity']}\n"
                f"       🏠 Xona: {g['room']}\n\n"
            )

    return text, inline_kb


@report_router.callback_query(F.data.startswith("wgp_"))
async def show_waiting_group_detail_handler(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    idx = int(call.data.split("_")[-1])
    await state.set_state(ReportStates.waiting_for_report_choice)
    await _show_waiting_group_detail(call, state, idx)


async def _show_waiting_group_detail(call: types.CallbackQuery, state: FSMContext, idx: int):
    state_data = await state.get_data()
    groups = state_data.get("waiting_groups", [])

    if idx >= len(groups):
        await call.message.edit_text("⚠️ Guruh topilmadi.")
        return

    g = groups[idx]
    all_comments = await get_all_comments()
    comment = all_comments.get(g["name"], "")

    detail = f"⏳ <b>{g['name']} — {g['level']}</b>\n\n"
    detail += f"👨🏻‍🏫 {g['teacher']}\n"
    detail += f"👥 {g['active']} + {g['frozen']} / {'?' if g['capacity'] == 0 else g['capacity']}\n"
    detail += f"🏠 Xona: {g['room']}\n"

    if comment:
        detail += f"\n📝 <b>Izoh:</b> {comment}\n"

    kb = []
    if comment:
        kb.append([
            types.InlineKeyboardButton(text="✏️ Tahrirlash", callback_data=f"wcmt_e_{idx}"),
            types.InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"wcmt_d_{idx}"),
        ])
    else:
        kb.append([
            types.InlineKeyboardButton(text="➕ Izoh qo'shish", callback_data=f"wcmt_a_{idx}"),
        ])
    kb.append([
        types.InlineKeyboardButton(text="⬅️ Waiting Groups", callback_data="wg_back"),
    ])

    await state.update_data(comment_group_name=g["name"], comment_group_idx=idx)

    await call.message.edit_text(
        detail,
        parse_mode="HTML",
        reply_markup=types.InlineKeyboardMarkup(inline_keyboard=kb),
    )


@report_router.callback_query(F.data == "wg_back")
async def back_to_waiting_groups(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    state_data = await state.get_data()
    groups = state_data.get("waiting_groups", [])
    all_comments = await get_all_comments()
    report, inline_kb = _render_waiting_groups_report(groups, all_comments)
    await state.set_state(ReportStates.waiting_for_report_choice)
    await call.message.edit_text(report, parse_mode="HTML", reply_markup=inline_kb)


@report_router.callback_query(F.data.startswith("wcmt_a_"))
@report_router.callback_query(F.data.startswith("wcmt_e_"))
async def start_waiting_comment_input(call: types.CallbackQuery, state: FSMContext):
    """➕ Izoh qo'shish yoki ✏️ tahrirlash"""
    await call.answer()
    idx = int(call.data.split("_")[-1])
    state_data = await state.get_data()
    groups = state_data.get("waiting_groups", [])

    if idx >= len(groups):
        await call.message.edit_text("⚠️ Guruh topilmadi.")
        return

    g = groups[idx]
    group_name = g["name"]
    all_comments = await get_all_comments()
    old_comment = all_comments.get(group_name, "")

    await state.update_data(comment_group_name=group_name, comment_group_idx=idx)
    await state.set_state(ReportStates.waiting_for_comment_input)

    hint = f"✏️ <b>{group_name}</b>\n\n"
    if old_comment:
        hint += f"📝 Joriy izoh: {old_comment}\n\n"
    hint += "Yangi izoh matnini kiriting (yoki ❌ Bekor qilish):"

    await call.message.edit_text(
        hint,
        parse_mode="HTML",
        reply_markup=types.InlineKeyboardMarkup(inline_keyboard=[
            [types.InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"wcmt_cancel_{idx}")]
        ]),
    )


@report_router.callback_query(F.data.startswith("wcmt_cancel_"))
async def cancel_waiting_comment(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    idx = int(call.data.split("_")[-1])
    await _show_waiting_group_detail(call, state, idx)


@report_router.callback_query(F.data.startswith("wcmt_d_"))
async def delete_waiting_comment(call: types.CallbackQuery, state: FSMContext):
    """Izohni o'chirish"""
    await call.answer()
    idx = int(call.data.split("_")[-1])
    state_data = await state.get_data()
    groups = state_data.get("waiting_groups", [])

    if idx >= len(groups):
        await call.message.edit_text("⚠️ Guruh topilmadi.")
        return

    group_name = groups[idx]["name"]
    await delete_comment(group_name)
    await _show_waiting_group_detail(call, state, idx)


# ================= USTOZ BO'YICHA CALLBACK HANDLERLAR =================
@report_router.message(ReportStates.waiting_for_teacher_choice, F.text == "📊 Finishing Groups")
async def switch_to_problematic(message: types.Message, state: FSMContext):
    await state.set_state(ReportStates.waiting_for_report_choice)
    groups = await asyncio.to_thread(get_all_groups)
    finishing = [g for g in groups if 0 <= g["days_left"] <= 14]
    if not finishing:
        await message.answer("📭 Finishing groups topilmadi.")
        return
    text = f"📋 <b>Finishing Groups</b> | Jami: {len(finishing)}\n\n"
    for g in sorted(finishing, key=lambda x: x["days_left"]):
        status = "⚪" if g["status"] == 1 else "🟢"
        text += f"{status} {g['name']} — {g['course']} | {g['teacher']}\n"
        text += f"    Tugash: {g['end']} ({g['days_left']} kun)\n\n"
    await message.answer(text, parse_mode="HTML")


@report_router.message(ReportStates.waiting_for_teacher_choice, F.text == "👨🏻‍🏫 Ustoz bo'yicha")
async def refresh_teachers(message: types.Message, state: FSMContext):
    await show_teachers_list(message, state)


@report_router.message(ReportStates.waiting_for_teacher_choice, F.text == "🏠 Bosh sahifa")
async def teacher_choice_home(message: types.Message, state: FSMContext):
    await state.clear()
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    from Keyboards.main_menu import get_main_menu
    await message.answer("🏠 Asosiy menyuga qaytdingiz.", reply_markup=get_main_menu(role))


@report_router.callback_query(ReportStates.waiting_for_teacher_choice, F.data == "teachergroups_cancel")
async def cancel_teacher_groups(call: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.delete()
    role = _USERS_ROLES.get(str(call.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    from Keyboards.main_menu import get_main_menu
    await call.message.answer("🏠 Asosiy menyuga qaytdingiz.", reply_markup=get_main_menu(role))
    await call.answer()


@report_router.callback_query(ReportStates.waiting_for_teacher_choice, F.data == "teachergroups_back")
async def back_to_teachers_list(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    teachers = get_unique_teachers()
    if not teachers:
        await call.message.edit_text("📭 LMSda o'qituvchilar topilmadi.")
        return

    inline_kb = []
    for t in teachers:
        inline_kb.append([types.InlineKeyboardButton(
            text=f"👨🏻‍🏫 {t}",
            callback_data=f"teachergroups_{t}",
        )])
    inline_kb.append([types.InlineKeyboardButton(
        text="🏠 Bosh sahifa",
        callback_data="teachergroups_cancel",
    )])

    await call.message.edit_text(
        text=f"👨🏻‍🏫 <b>Ustozni tanlang:</b>\n\nJami {len(teachers)} ta o'qituvchi",
        parse_mode="HTML",
        reply_markup=types.InlineKeyboardMarkup(inline_keyboard=inline_kb),
    )


@report_router.callback_query(ReportStates.waiting_for_teacher_choice, F.data.startswith("teachergroups_"))
async def show_teacher_groups(call: types.CallbackQuery, state: FSMContext):
    teacher_name = call.data.replace("teachergroups_", "")
    await call.answer()
    await call.message.edit_text(f"⏳ <b>{teacher_name}</b> guruhlari yuklanmoqda...", parse_mode="HTML")

    groups = await asyncio.to_thread(get_all_groups)
    teacher_groups = [g for g in groups if g["teacher"] == teacher_name]

    if not teacher_groups:
        await call.message.edit_text(
            text=f"👨🏻‍🏫 <b>{teacher_name}</b>\n\n📭 Hozirda faol guruhlari topilmadi.",
            parse_mode="HTML",
        )
        return

    text = f"👨🏻‍🏫 <b>{teacher_name}</b> — {len(teacher_groups)} ta guruh\n\n"
    for g in sorted(teacher_groups, key=lambda x: x["start"]):
        days_str = f"({g['days_left']} kun qoldi)" if g["days_left"] != 999 else ""
        text += f"📚 {g['name']} — {g['course']}\n"
        text += f"   📅 {g['start']} → {g['end']} {days_str}\n\n"

    if len(text) > 4096:
        text = text[:4000] + "\n\n... (davomi bor)"

    await call.message.edit_text(
        text,
        parse_mode="HTML",
        reply_markup=types.InlineKeyboardMarkup(inline_keyboard=[
            [types.InlineKeyboardButton(text="⬅️ Ustozlar ro'yxati", callback_data="teachergroups_back")],
            [types.InlineKeyboardButton(text="🏠 Bosh sahifa", callback_data="teachergroups_cancel")],
        ]),
    )


# ================= CASHBOX CALLBACK HANDLER =================
@report_router.callback_query(ReportStates.waiting_for_cashbox_detail, F.data.startswith("cb_"))
async def cashbox_callback_handler(call: types.CallbackQuery, state: FSMContext):
    """Cashbox tanlanganda yoki orqaga qaytganda"""
    await call.answer()
    data = call.data

    if data == "cb_home":
        await state.clear()
        await call.message.delete()
        role = _USERS_ROLES.get(str(call.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
        from Keyboards.main_menu import get_main_menu
        await call.message.answer("🏠 Asosiy menyuga qaytdingiz.", reply_markup=get_main_menu(role))
        return

    if data == "cb_back":
        msg = call.message
        try:
            s = _get_lms_session()
            r = s.get(f"{LMS_BASE}/admin/cashboxes", timeout=15)
            match = re.search(r'data-page="([^"]*)"', r.text)
            if not match:
                await msg.edit_text("⚠️ Cashbox ma'lumotlari topilmadi.")
                return

            dp = json.loads(unescape(match.group(1)))
            cashboxes = dp["props"]["cashboxes"]
            drujba_cbs = [cb for cb in cashboxes if cb.get("branch", {}).get("en") == "Drujba filial"]

            text = "🏦 <b>DRUJBA — CASHBOXLAR</b>\n\n"
            from_keyboard = []
            for idx, cb in enumerate(drujba_cbs, 1):
                bal = cb.get("balance", {}) if isinstance(cb.get("balance"), dict) else {}
                total = sum(float(v or 0) for v in bal.values())
                text += f"{idx}. <b>{cb['name']}</b> — 💰 {int(total)} so'm\n"
                from_keyboard.append([types.InlineKeyboardButton(text=f"{cb['name']} — {int(total)} so'm", callback_data=f"cb_{cb['id']}")])

            from_keyboard.append([types.InlineKeyboardButton(text="🏠 Bosh sahifa", callback_data="cb_home")])
            inline_kb = types.InlineKeyboardMarkup(inline_keyboard=from_keyboard)

            await msg.edit_text(text, parse_mode="HTML", reply_markup=inline_kb)
        except Exception as e:
            await msg.edit_text(f"⚠️ Xatolik: {e}")
        return

    # Cashbox detail
    cb_id = data.replace("cb_", "")
    msg = call.message

    try:
        s = _get_lms_session()
        r = s.get(f"{LMS_BASE}/admin/cashboxes", timeout=15)
        match = re.search(r'data-page="([^"]*)"', r.text)
        if not match:
            await msg.edit_text("⚠️ Cashbox ma'lumotlari topilmadi.")
            return

        dp = json.loads(unescape(match.group(1)))
        cashboxes = dp["props"]["cashboxes"]
        cb = next((c for c in cashboxes if str(c["id"]) == cb_id), None)

        if not cb:
            await msg.edit_text("⚠️ Cashbox topilmadi.")
            return

        bal = cb.get("balance", {})
        if not isinstance(bal, dict):
            bal = {}

        total = sum(float(v or 0) for v in bal.values())

        text = f"🏦 <b>{cb['name']}</b>\n\n"
        text += f"💰 <b>Jami:</b> {int(total)} so'm\n"
        text += f"━━━━━━━━━━━━━━━━\n\n"

        cash_v = float(bal.get('cash', 0) or 0)
        terminal_v = float(bal.get('terminal', 0) or 0)
        qr_v = float(bal.get('qrcode', 0) or 0)
        mchj_v = float(bal.get('llcaccounts', 0) or 0)

        if cash_v: text += f"💵 <b>Naqd:</b> {int(cash_v)} so'm\n"
        if terminal_v: text += f"💳 <b>Terminal:</b> {int(terminal_v)} so'm\n"
        if qr_v: text += f"📱 <b>QR:</b> {int(qr_v)} so'm\n"
        if mchj_v: text += f"🏛 <b>MCHJ hisob raqamlar:</b> {int(mchj_v)} so'm\n"

        if total == 0:
            text += "📭 Bu cashboxda pul mavjud emas.\n"

        inline_kb = types.InlineKeyboardMarkup(
            inline_keyboard=[
                [types.InlineKeyboardButton(text="🏦 Ortga (Cashboxlar)", callback_data="cb_back")]
            ]
        )

        await msg.edit_text(text, parse_mode="HTML", reply_markup=inline_kb)

    except Exception as e:
        logger.error(f"Cashbox detail error: {e}")
        await msg.edit_text(f"⚠️ Xatolik: {e}")


# ================= USTOZLARNI BOSHQARISH =================

@report_router.message(ReportStates.waiting_for_report_choice, F.text == "👨🏻‍🏫 Ustozlarni boshqarish")
async def manage_teachers(message: types.Message, state: FSMContext):
    """Ustozlar ro'yxatini ko'rish, qo'shish, olib tashlash"""
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    if role != "Owner":
        await message.answer("⚠️ Bu buyruq faqat Owner uchun!")
        return

    teachers = _load_teachers()
    text = "👨🏻‍🏫 <b>DRUJBA USTOZLARI RO'YXATI</b>\n\n"
    for idx, t in enumerate(teachers, 1):
        text += f"{idx}. {t}\n"

    from_keyboard = [
        [types.InlineKeyboardButton(text="➕ Yangi ustoz qo'shish", callback_data="teacher_add")],
        [types.InlineKeyboardButton(text="➖ Ustozni olib tashlash", callback_data="teacher_remove")],
        [types.InlineKeyboardButton(text="🏠 Bosh sahifa", callback_data="teacher_home")],
    ]
    inline_kb = types.InlineKeyboardMarkup(inline_keyboard=from_keyboard)

    msg = await message.answer(text, parse_mode="HTML", reply_markup=inline_kb)
    await state.set_state(ReportStates.waiting_for_teacher_choice)


@report_router.callback_query(F.data == "teacher_home")
async def teacher_home(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    await state.clear()
    await call.message.delete()
    role = _USERS_ROLES.get(str(call.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    from Keyboards.main_menu import get_main_menu
    await call.message.answer("🏠 Asosiy menyuga qaytdingiz.", reply_markup=get_main_menu(role))


@report_router.callback_query(F.data == "teacher_add")
async def teacher_add_prompt(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    await call.message.edit_text(
        "✏️ Yangi ustozning to'liq ismini kiriting:\n\n"
        "Masalan: <code>Bobur Aliyev</code>\n\n"
        "Bekor qilish uchun /cancel yozing.",
        parse_mode="HTML"
    )
    await state.set_state(ReportStates.waiting_for_teacher_name_add)


@report_router.message(ReportStates.waiting_for_teacher_name_add)
async def teacher_add_confirm(message: types.Message, state: FSMContext):
    name = message.text.strip()
    if not name or len(name) < 3:
        await message.answer("⚠️ Ism juda qisqa. Iltimos, to'liq ismni kiriting (kamida 3 harf).")
        return

    teachers = _load_teachers()

    # Takrorlanmasligini tekshirish
    for t in teachers:
        if name.lower() in t.lower() or t.lower() in name.lower():
            await message.answer(
                f"⚠️ <b>{name}</b> ro'yxatda allaqachon mavjud yoki o'xshash ism bor:\n"
                f"   → {t}\n\n"
                f"Qayta urinib ko'ring yoki /cancel yozing.",
                parse_mode="HTML"
            )
            return

    teachers.append(name)
    _save_teachers(teachers)

    text = f"✅ <b>{name}</b> ro'yxatga qo'shildi!\n\n"
    text += "👨🏻‍🏫 <b>Joriy ro'yxat:</b>\n"
    for idx, t in enumerate(teachers, 1):
        text += f"{idx}. {t}\n"

    from_keyboard = [
        [types.InlineKeyboardButton(text="➕ Yana qo'shish", callback_data="teacher_add")],
        [types.InlineKeyboardButton(text="🏁 Tugatish", callback_data="teacher_home")],
    ]
    inline_kb = types.InlineKeyboardMarkup(inline_keyboard=from_keyboard)
    await message.answer(text, parse_mode="HTML", reply_markup=inline_kb)
    await state.set_state(ReportStates.waiting_for_teacher_choice)


@report_router.callback_query(F.data == "teacher_remove")
async def teacher_remove_prompt(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    teachers = _load_teachers()

    text = "🗑 <b>Olib tashlash uchun ustozni tanlang:</b>\n\n"
    from_keyboard = []
    for idx, t in enumerate(teachers, 1):
        from_keyboard.append([
            types.InlineKeyboardButton(text=f"{idx}. {t}", callback_data=f"tremove_{idx}")
        ])
    from_keyboard.append([types.InlineKeyboardButton(text="🔙 Ortga", callback_data="teacher_back")])

    inline_kb = types.InlineKeyboardMarkup(inline_keyboard=from_keyboard)
    await call.message.edit_text(text, parse_mode="HTML", reply_markup=inline_kb)


@report_router.callback_query(F.data.startswith("tremove_"))
async def teacher_remove_confirm(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    idx = int(call.data.replace("tremove_", "")) - 1
    teachers = _load_teachers()

    if idx < 0 or idx >= len(teachers):
        await call.message.edit_text("⚠️ Xatolik: bunday ustoz topilmadi.")
        return

    removed = teachers.pop(idx)
    _save_teachers(teachers)

    text = f"✅ <b>{removed}</b> ro'yxatdan olib tashlandi!\n\n"
    text += "👨🏻‍🏫 <b>Yangilangan ro'yxat:</b>\n"
    for i, t in enumerate(teachers, 1):
        text += f"{i}. {t}\n"

    from_keyboard = [
        [types.InlineKeyboardButton(text="➖ Yana olib tashlash", callback_data="teacher_remove")],
        [types.InlineKeyboardButton(text="🏁 Tugatish", callback_data="teacher_home")],
    ]
    inline_kb = types.InlineKeyboardMarkup(inline_keyboard=from_keyboard)
    await call.message.edit_text(text, parse_mode="HTML", reply_markup=inline_kb)


@report_router.callback_query(F.data == "teacher_back")
async def teacher_back(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    await manage_teachers(call.message, state)


# ================= DEBTORS SUBMENU =================

def _debt_percent_coeff(percent: float) -> float:
    """Qarzdorlik foiziga qarab Kassir maoshi koeffitsiyentini qaytaradi."""
    if percent <= 0:
        return 3.0
    if percent < 2.0:
        return 2.0
    if percent < 5.0:
        return 1.8
    if percent < 7.0:
        return 1.7
    if percent < 10.0:
        return 1.6
    if percent < 15.0:
        return 1.5
    if percent < 20.0:
        return 1.4
    return 1.2

@report_router.message(ReportStates.waiting_for_report_choice, F.text == "💲 Debtors")
async def debtors_submenu(message: types.Message, state: FSMContext):
    """Debtors submanyusi: '🟢 Active Debs' va 'Debs ％' tugmalari."""
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", None) if _USERS_ROLES else None
    if role != "Kassir" and not await is_admin(message.from_user.id):
        return
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    if role == "Kassir":
        # Kassirga faqat 'Debs ％' (foiz) ko'rinadi — '🟢 Active Debs' (Sheets ga yozuvchi) ko'rinmaydi
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Debs ％")],
                [KeyboardButton(text="⬅️ Ortga")],
            ],
            resize_keyboard=True,
        )
    else:
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="🟢 Active Debs")],
                [KeyboardButton(text="Debs ％")],
                [KeyboardButton(text="⬅️ Ortga")],
            ],
            resize_keyboard=True,
        )
    await message.answer("💲 <b>Debtors</b>\n\nQarzdor talabalar ro'yxati va qarzdorlik foizi uchun quyidagilarni bosing:", parse_mode="HTML", reply_markup=kb)


@report_router.message(ReportStates.waiting_for_report_choice, F.text == "🟢 Active Debs")
async def active_debs_handler(message: types.Message, state: FSMContext):
    """Google Sheets ga barcha qarzdorlarni yozadi (deb-eksport)."""
    if not await is_admin(message.from_user.id):
        return
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🟢 Active Debs")],
            [KeyboardButton(text="Debs ％")],
            [KeyboardButton(text="⬅️ Ortga")],
        ],
        resize_keyboard=True,
    )
    await message.answer("⏳ <b>Qarzdorlarni yuklab olmoqdaman...</b>\nBu bir necha daqiqa olishi mumkin.", parse_mode="HTML")

    try:
        import asyncio
        from utils.debtors_export import run_export

        def _do():
            return run_export()

        info = await asyncio.to_thread(_do)
        total = info["total"]
        rows = info["rows_written"]
        if total == 0:
            text = ("⚠️ <b>Ma'lumot topilmadi.</b>\n\n"
                    "LMS dan qarzdorlar olinmadi. Ehtimol LMS paroli to'g'ri emas yoki boshqa xatolik.\n"
                    "Admin': LMS_KEY muhit o'zgaruvchisini <b>Mahmudov02</b> ga o'zgartiring.")
        else:
            text = (f"✅ <b>Done!</b> Google Sheets ga yozildi.\n\n"
                    f"👥 Jami: <b>{total}</b> ta qarzdor yozildi.\n"
                    f"🔗 <a href='https://docs.google.com/spreadsheets/d/1PpGWObeppzsSkaYgGz0fRYP_3zk-3YuxBOXStrn_PCc'>Google Sheets ni ochish</a>")
        await message.answer(text, parse_mode="HTML", reply_markup=kb, disable_web_page_preview=True)
    except Exception as e:
        logger.exception("Debtors export xatosi")
        await message.answer(f"❌ <b>Xatolik:</b> {e}", parse_mode="HTML", reply_markup=kb)


@report_router.message(ReportStates.waiting_for_report_choice, F.text == "Debs ％")
async def debt_percent_handler(message: types.Message, state: FSMContext):
    """LMS dan qarzdorlik foizini hisoblab, botda ko'rsatadi."""
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", None) if _USERS_ROLES else None
    if role != "Kassir" and not await is_admin(message.from_user.id):
        return
    from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
    if role == "Kassir":
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Debs ％")],
                [KeyboardButton(text="⬅️ Ortga")],
            ],
            resize_keyboard=True,
        )
    else:
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="🟢 Active Debs")],
                [KeyboardButton(text="Debs ％")],
                [KeyboardButton(text="⬅️ Ortga")],
            ],
            resize_keyboard=True,
        )
    await message.answer("⏳ <b>Qarzdorlik foizini hisoblamoqdaman...</b>\nBu bir necha daqiqa olishi mumkin.", parse_mode="HTML")

    try:
        import asyncio
        from datetime import datetime as _dt
        from utils.debtors_export import get_debt_percent_data

        def _do():
            return get_debt_percent_data()

        d = await asyncio.to_thread(_do)

        active_total = d["active_total"]
        archive_total = d["archive_total"]
        active_debt = d["active_debt"]
        archive_debt = d["archive_debt"]
        all_total = active_total + archive_total
        all_debt = active_debt + archive_debt
        percent = (all_debt / all_total * 100) if all_total else 0

        today = _dt.now()
        date_str = today.strftime("%d.%m.%Y")
        percent_txt = f"{percent:.1f}".rstrip("0").rstrip(".") if percent == int(percent) else f"{percent:.1f}"

        # Koeffitsiyentni foizga qarab hisoblash (Kassir maoshi uchun)
        coeff = _debt_percent_coeff(percent)
        coeff_txt = f"{coeff:g}"

        text = (
            "📊 <b>Qarzdorlik foizi</b>\n\n"
            f"🗓 Davr: {date_str}\n\n"
            f"👥 Umumiy: {all_total:.0f} ta\n"
            f"✅ Aktiv: {active_total:.0f}\n"
            f"🗄 Arxiv: {archive_total:.0f}\n\n"
            f"💰 Qarzdor: {all_debt:.0f} ta\n"
            f"⚠️ Aktiv: {active_debt:.0f}\n"
            f"🗄 Arxiv: {archive_debt:.0f}\n\n"
            f"📈 Foiz: {percent_txt}%  (×{coeff_txt})"
        )
        await message.answer(text, parse_mode="HTML", reply_markup=kb)

        # Maosh hisoblash taklifi
        from aiogram.types import ReplyKeyboardMarkup as _RKM, KeyboardButton as _KB
        choice_kb = _RKM(
            keyboard=[
                [_KB(text="Istayman 🙃"), _KB(text="Istamayman")],
            ],
            resize_keyboard=True,
        )
        # Foiz koeffitsiyentini state'da saqlaymiz
        await state.update_data(
            debt_coeff=coeff,
            debt_percent=percent_txt,
            debt_date=date_str,
        )
        await state.set_state(ReportStates.waiting_for_salary_choice)
        await message.answer(
            "Istasangiz men sizga umumiy maoshingizni hisoblab beraman 🙃",
            reply_markup=choice_kb,
        )
    except Exception as e:
        logger.exception("Debt percent xatosi")
        await message.answer(f"❌ <b>Xatolik:</b> {e}", parse_mode="HTML", reply_markup=kb)


# ================= QARZDORLIK FOIZI -> MAOSH HISOBLASH =================

@report_router.message(ReportStates.waiting_for_salary_choice, F.text == "Istamayman")
async def salary_decline(message: types.Message, state: FSMContext):
    """Foydalanuvchi maosh hisoblamaslikni tanlasa — bosh menyuga qaytadi."""
    await state.clear()
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    from Keyboards.main_menu import get_main_menu
    await message.answer("🏠 <b>Bosh sahifa</b>", parse_mode="HTML", reply_markup=get_main_menu(role))


@report_router.message(ReportStates.waiting_for_salary_choice, F.text == "Istayman 🙃")
async def salary_start(message: types.Message, state: FSMContext):
    """Maosh hisoblash boshlandi — kunlik soatlar so'raladi."""
    await state.set_state(ReportStates.waiting_for_salary_hours)
    await message.answer(
        "1️⃣ Siz bir kunda necha soat ishlaysiz?\n\n"
        "<i>Faqat raqam kiriting (masalan: 8)</i>",
        parse_mode="HTML",
        reply_markup=types.ReplyKeyboardRemove(),
    )


@report_router.message(ReportStates.waiting_for_salary_hours)
async def salary_hours_input(message: types.Message, state: FSMContext):
    """Kunlik soatlarni qabul qiladi — faqat raqam."""
    txt = message.text.strip()
    if not txt.isdigit():
        await message.answer("⚠️ Iltimos, faqat raqam kiriting (masalan: 8).")
        return
    hours = int(txt)
    if hours <= 0:
        await message.answer("⚠️ Soat soni 0 dan katta bo'lishi kerak.")
        return
    await state.update_data(salary_hours=hours)
    await state.set_state(ReportStates.waiting_for_salary_days)
    await message.answer(
        "2️⃣ Hisoblamoqchi bo'lgan oy davomida necha kun ishladingiz?\n\n"
        "<i>Yakshanba, bayram kunlari va ishlamagan kunlaringizdan tashqari. Faqat raqam kiriting.</i>",
        parse_mode="HTML",
    )


@report_router.message(ReportStates.waiting_for_salary_days)
async def salary_days_input(message: types.Message, state: FSMContext):
    """Ishlangan kunlarni qabul qiladi — faqat raqam."""
    txt = message.text.strip()
    if not txt.isdigit():
        await message.answer("⚠️ Iltimos, faqat raqam kiriting.")
        return
    days = int(txt)
    if days <= 0:
        await message.answer("⚠️ Kunlar soni 0 dan katta bo'lishi kerak.")
        return
    await state.update_data(salary_days=days)
    await state.set_state(ReportStates.waiting_for_salary_cover)

    from aiogram.types import ReplyKeyboardMarkup as _RKM, KeyboardButton as _KB
    cover_kb = _RKM(
        keyboard=[
            [_KB(text="Cover qilmadim")],
        ],
        resize_keyboard=True,
    )
    await message.answer(
        "3️⃣ Cover qilgan soatlaringizni yig'indisini kiriting.\n\n"
        "<i>Raqam kiriting (masalan: 14). Agar cover bo'lmasa — 'Cover qilmadim' tugmasini bosing.</i>",
        parse_mode="HTML",
        reply_markup=cover_kb,
    )


@report_router.message(ReportStates.waiting_for_salary_cover, F.text == "Cover qilmadim")
async def salary_cover_none(message: types.Message, state: FSMContext):
    """Cover bo'lmasa — 0 cover bilan hisoblaydi."""
    await _compute_salary(message, state, cover_hours=0)


@report_router.message(ReportStates.waiting_for_salary_cover)
async def salary_cover_input(message: types.Message, state: FSMContext):
    """Cover soatlarini qabul qiladi va maoshni hisoblaydi."""
    txt = message.text.strip().replace(" ", "")
    if not txt.lstrip("-").isdigit():
        await message.answer("⚠️ Iltimos, faqat raqam kiriting (yoki 'Cover qilmadim' tugmasini bosing).")
        return
    cover = int(txt)
    if cover < 0:
        await message.answer("⚠️ Cover soati manfiy bo'lolmaydi.")
        return
    await _compute_salary(message, state, cover_hours=cover)


async def _compute_salary(message, state, cover_hours: int):
    """Kassir fiks maoshini hisoblab ko'rsatadi."""
    data = await state.get_data()
    hours = data.get("salary_hours", 0)
    days = data.get("salary_days", 0)
    coeff = data.get("debt_coeff", 1.0)
    percent_txt = data.get("debt_percent", "?")
    date_str = data.get("debt_date", "?")

    # Fiks hisoblash: birinchi 8 soat 15,000/soat, 8 dan oshgani 20,000/soat
    base_hours = 8
    rate_base = 15000
    rate_overtime = 20000
    if hours <= base_hours:
        daily = hours * rate_base
    else:
        daily = base_hours * rate_base + (hours - base_hours) * rate_overtime

    monthly_fix = daily * days
    cover_amount = cover_hours * 15000  # cover har doim 15,000 dan (20,000 ga o'tmaydi)
    total_before = monthly_fix + cover_amount
    final = total_before * coeff

    def _fmt(n):
        return f"{int(round(n)):,}".replace(",", " ")

    # Hisobni tushuntirish
    if hours <= base_hours:
        daily_explain = f"{hours} soat × 15,000 = {_fmt(daily)}"
    else:
        extra = hours - base_hours
        daily_explain = (f"8 soat × 15,000 + {extra} soat × 20,000 = "
                         f"{_fmt(base_hours*rate_base)} + {_fmt(extra*rate_overtime)} = {_fmt(daily)}")

    text = (
        "💼 <b>Umumiy maosh (Kassir)</b>\n\n"
        f"🗓 Davr: {date_str}\n"
        f"📈 Qarzdorlik foizi: {percent_txt}%  (×{coeff:g})\n\n"
        f"⏱ <b>Kunlik:</b> {daily_explain}\n"
        f"📆 Ishlangan kunlar: {days} kun\n"
        f"💰 <b>Oylik fiks:</b> {_fmt(daily)} × {days} = {_fmt(monthly_fix)}\n"
        f"🔁 Cover: {cover_hours} soat × 15,000 = {_fmt(cover_amount)}\n"
        f"➡️ <b>Jami (fiks + cover):</b> {_fmt(total_before)}\n\n"
        f"🎯 <b>Yakuniy maosh:</b>\n"
        f"{_fmt(total_before)} × {coeff:g} = <b>{_fmt(final)} so'm</b>"
    )

    await state.clear()
    role = _USERS_ROLES.get(str(message.from_user.id), {}).get("role", "Owner") if _USERS_ROLES else "Owner"
    from Keyboards.main_menu import get_main_menu
    await message.answer(text, parse_mode="HTML", reply_markup=get_main_menu(role))
